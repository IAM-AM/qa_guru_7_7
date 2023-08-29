from openpyxl import load_workbook
import os
from conftest import RESOURCE_ROOT_PATH


# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx_file():
    file_path = os.path.join(RESOURCE_ROOT_PATH, "file_example_XLSX_50.xlsx")
    workbook = load_workbook(file_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    assert os.path.isfile(file_path)
    assert sheet.cell(row=3, column=2).value == "Mara"
    assert file_path.endswith('.xlsx')

    os.remove(file_path)
